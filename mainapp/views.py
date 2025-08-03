# products/views.py

import json
import traceback
from django.utils import timezone
from datetime import datetime, timedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.http import JsonResponse, Http404  # To send JSON responses for AJAX
from django.shortcuts import redirect, render
from django.utils.decorators import method_decorator
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt  # For AJAX, consider CSRF tokens properly in production
from .models import OTP  # Import the OTP model
from .models import Cart, CartItem, Product, Category, Order
from django.db.models import Count
import uuid
import requests
from django.utils.timesince import timesince
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Product
from rest_framework import viewsets
from .serializers import ProductSerializer, CategorySerializer, OrderSerializer, UserSerializer, ProfileSerializer, AddressSerializer
from decimal import Decimal
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from .models import Profile, Address
from django.shortcuts import render 
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.core.mail import send_mail
from django.conf import settings
import json
from .models import Order

# app_name/views.py
from django.http import JsonResponse
from .models import StoreSettings

def get_shipping_cost(request):
    try:
        settings = StoreSettings.objects.first()
        if settings is None:
            return JsonResponse({'error': 'Pengaturan toko belum dikonfigurasi'}, status=404)
        
        return JsonResponse({
            'shipping_cost': settings.shipping_cost,
            'free_shipping_min': settings.free_shipping_min
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Existing views (keep them as they are)
def admin_login_view(request):
    return render(request, 'admin_login.html', {})

def admin_dashboard_view(request):
    return render(request, 'admin_dashboard.html', {})

def dashboard(request):
    return render(request, 'dashboard.html')

def cart(request):
    return render(request, 'cart.html')

def checkout(request):
    return render(request, 'checkout.html')

def profile(request):
    return render(request, 'profile.html')

def order(request):
    return render(request, 'order_success.html')

def order_history(request):
    return render(request, 'order_history.html')

def privacy_policy(request):
    return render(request, 'privacy_policy.html')

def help_faq(request):
    return render(request, 'help_faq.html')

def tutorial(request):
    return render(request, 'tutorial.html')

def return_req(request):
    return render(request, 'return_req.html')

def payment_accounts_view (request):
    return render(request, 'payment_accounts.html')

def confirmation_payment_view (request):
    return render(request, 'confirm_payment.html')

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all().select_related('user').prefetch_related('items')
    serializer_class = OrderSerializer

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Order
from .serializers import OrderSerializer, OrderItemSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db.models import F, Sum
from django.db.models.functions import TruncMonth
from .models import OrderItem
import calendar

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import StoreSettings
from .serializers import StoreSettingsSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

@api_view(['GET', 'POST'])
def store_settings_view(request):
    # Get the single instance
    settings_instance, _ = StoreSettings.objects.get_or_create(pk=1)

    if request.method == 'GET':
        serializer = StoreSettingsSerializer(settings_instance)
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = StoreSettingsSerializer(settings_instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Pengaturan berhasil disimpan'}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required
@csrf_exempt
def get_user_orders(request):
    if request.method == 'GET':
        orders = Order.objects.filter(user=request.user).order_by('-created_at')
        serializer = OrderSerializer(orders, many=True)
        return JsonResponse(serializer.data, safe=False)
    return JsonResponse({"error": "Method not allowed"}, status=405)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_order_items(request):
    order_id = request.GET.get('order_id')
    query = OrderItem.objects.filter(order__user=request.user)

    if order_id:
        query = query.filter(order__order_id=order_id)

    query = query.select_related('product', 'order')
    serializer = OrderItemSerializer(query, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def sales_analytics(request):
    monthly_sales = (
        OrderItem.objects
        .annotate(month=TruncMonth('order__created_at'))
        .values('month')
        .annotate(total=Sum(F('price') * F('quantity')))
        .order_by('month')
    )

    result = {
        calendar.month_abbr[item['month'].month]: float(item['total']) if item['total'] else 0
        for item in monthly_sales if item['month']
    }

    return Response(result)



@api_view(['GET'])
def dashboard_summary(request):
    total_penjualan = sum(order.total_price for order in Order.objects.all())
    pesanan_baru = Order.objects.filter(payment_status='pending').count()
    jumlah_produk = Product.objects.count()
    pelanggan_baru = User.objects.count()

    aktivitas = []
    recent_orders = Order.objects.order_by('-created_at')[:5]
    for order in recent_orders:
        waktu = timesince(order.created_at).split(',')[0] + " yang lalu"
        aktivitas.append(f"Order {order.order_id} dibuat oleh {order.user.username if order.user else 'Guest'} ({waktu})")

    return Response({
        "total_penjualan": total_penjualan,
        "pesanan_baru": pesanan_baru,
        "jumlah_produk": jumlah_produk,
        "pelanggan_baru": pelanggan_baru,
        "aktivitas": aktivitas,
    })

def get_product_detail(request, product_id):
    try:
        product = Product.objects.select_related('category').get(id=product_id)
        return JsonResponse({
            "id": product.id,
            "name": product.name,
            "category_id": product.category.id if product.category else None,
            "category": product.category.name if product.category else "",
            "stock": product.stock,
            "price": product.price,
            "description": product.description,
        })
    except Product.DoesNotExist:
        raise Http404("Produk tidak ditemukan")
    
    # views.py
@api_view(['POST'])
def update_product(request, pk):
    try:
        product = Product.objects.get(pk=pk)
    except Product.DoesNotExist:
        return Response({'success': False, 'message': 'Produk tidak ditemukan'}, status=404)

    product.name = request.POST.get('name')
    product.category_id = request.POST.get('category')
    product.price = request.POST.get('price')
    product.stock = request.POST.get('stock')
    product.description = request.POST.get('description')

    if 'image' in request.FILES:
        product.image = request.FILES['image']

    product.save()

    return Response({'success': True, 'message': 'Produk berhasil diperbarui'})

@api_view(['DELETE'])
def delete_product(request, pk):
    try:
        product = Product.objects.get(pk=pk)
        product.delete()
        return Response({'success': True, 'message': 'Produk berhasil dihapus'})
    except Product.DoesNotExist:
        return Response({'success': False, 'message': 'Produk tidak ditemukan'}, status=404)
    

def get_category_detail(request, category_id):
    try:
        category = Category.objects.get(pk=category_id)
        return JsonResponse({
            "id": category.id,
            "name": category.name,
        })
    except Category.DoesNotExist:
        raise Http404("Kategori tidak ditemukan")
    
    # views.py
@api_view(['POST'])
def update_category(request, category_id):
    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return Response({'success': False, 'message': 'Kategori tidak ditemukan'}, status=404)

    category.name = request.POST.get('name')

    category.save()

    return Response({'success': True, 'message': 'Kategori berhasil diperbarui'})

@api_view(['DELETE'])
def delete_category(request, category_id):
    try:
        category = Category.objects.get(id=category_id)
        category.delete()
        return Response({'success': True, 'message': 'Kategori berhasil dihapus'})
    except Category.DoesNotExist:
        return Response({'success': False, 'message': 'Kategori tidak ditemukan'}, status=404)

@csrf_exempt
def midtrans_callback(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            order_id = data.get('order_id')
            transaction_status = data.get('transaction_status')

            # Cari order dari database
            order = Order.objects.get(order_id=order_id)
            order.payment_status = transaction_status
            order.save()

            # ✅ Jika transaksi berhasil, kirim email notifikasi ke user
            if transaction_status == 'settlement' and order.user and order.user.email:
                subject = 'Pembayaran Berhasil - Terusan Minimarket'
                message = (
                    f"Halo {order.user.username},\n\n"
                    f"Terima kasih! Pembayaran untuk pesanan #{order.order_id} telah berhasil.\n\n"
                    f"Detail Pesanan:\n"
                    f"Total: Rp {int(order.total_price):,}\n"
                    f"Status Pembayaran: {transaction_status}\n\n"
                    f"Pesananmu akan segera diproses.\n\n"
                    f"Salam,\nTerusan Minimarket"
                ) 
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,  # pastikan ini sudah di-set
                    [order.user.email],
                    fail_silently=False
                )
            
            return JsonResponse({'status': 'success'})

        except Order.DoesNotExist:
            return JsonResponse({'error': 'Order tidak ditemukan'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Only POST allowed'}, status=405)


from django.http import JsonResponse
from .models import Order
from django.views.decorators.http import require_GET

@require_GET
def order_details_api(request):
    order_id = request.GET.get('order_id')
    if not order_id:
        return JsonResponse({'error': 'order_id dibutuhkan'}, status=400)

    try:
        order = Order.objects.get(order_id=order_id)
        return JsonResponse({
            'order_id': order.order_id,
            'payment_method': order.payment_method,
            'shipping_address': order.shipping_address,
            'delivery_estimate': order.delivery_estimate,
            'total_price': float(order.total_price),
        })
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Pesanan tidak ditemukan'}, status=404)

@csrf_exempt
def create_transaction(request):
    if request.method != 'POST':
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        # --- Ambil data dari request body ---
        data = json.loads(request.body)
        amount = int(data.get("amount", 0))
        shipping_address = data.get("shipping_address", "")
        items_data = data.get("items", [])

        print("ITEMS RECEIVED:", items_data)

        # --- Validasi awal ---
        if amount <= 0 or not shipping_address or not items_data:
            return JsonResponse({"error": "Data tidak lengkap"}, status=400)

        # --- Siapkan user dan order_id ---
        user = request.user if request.user.is_authenticated else None
        order_id = f"TRSMK-{uuid.uuid4().hex[:10]}"

        # --- Buat Order ---
        order = Order.objects.create(
            user=user,
            order_id=order_id,
            payment_method="Midtrans",
            payment_status="pending",
            shipping_address=shipping_address,
            delivery_estimate="-",
            total_price=amount,
            created_at=timezone.now()
        )

        # --- Buat OrderItem dari items_data ---
        for item in items_data:
            product = Product.objects.get(id=item['product_id'])
            quantity = int(item['quantity'])

              # Kurangi stok produk
            if product.stock < quantity:
                return JsonResponse({"success": False, "message": f"Stok tidak cukup untuk {product.name}."}, status=400)

            product.stock -= quantity
            product.save()

            OrderItem.objects.create(
                order=order,
                product=product,
                quantity=item['quantity'],
                price=Decimal(str(item['price']))  # Penting: set harga
            )

        # --- Kosongkan keranjang jika user login ---
        if user:
            try:
                user_cart = Cart.objects.get(user=user)
                user_cart.items.all().delete()
            except Cart.DoesNotExist:
                pass  # Tidak masalah jika cart tidak ditemukan

        # --- Payload untuk Midtrans Snap ---
        payload = {
            "transaction_details": {
                "order_id": order_id,
                "gross_amount": amount
            },
            "credit_card": {"secure": True},
            "customer_details": {
                "first_name": user.username if user else "Guest",
                "email": user.email if user else "guest@example.com"
            }
        }

        # --- Kirim request ke Midtrans ---
        response = requests.post(
            "https://app.sandbox.midtrans.com/snap/v1/transactions",
            json=payload,
            auth=(settings.MIDTRANS_SERVER_KEY, '')
        )
        print(response.json())

        # --- Kirim response Midtrans ke client ---
        midtrans_response = response.json()
        midtrans_response["order_id"] = order_id
        return JsonResponse(midtrans_response)

    except Product.DoesNotExist:
        return JsonResponse({"error": "Produk tidak ditemukan"}, status=404)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@login_required
def create_cod_transaction(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = int(data.get("amount", 0))
            shipping_address = data.get("shipping_address", "")
            items_data = data.get("items", [])

            if amount <= 0 or not shipping_address or not items_data:
                return JsonResponse({"success": False, "message": "Data tidak lengkap"}, status=400)

            order_id = f"COD-{uuid.uuid4().hex[:8]}"
            user = request.user

            # Buat order
            order = Order.objects.create(
                user=user,
                order_id=order_id,
                payment_method="COD",
                payment_status="settlement",
                shipping_address=shipping_address,
                delivery_estimate="-",
                total_price=amount,
                created_at=timezone.now()
            )

            for item in items_data:
                product = Product.objects.get(id=item['product_id'])
                quantity = int(item['quantity'])

                # Kurangi stok produk
                if product.stock < quantity:
                    return JsonResponse({"success": False, "message": f"Stok tidak cukup untuk {product.name}."}, status=400)

                product.stock -= quantity
                product.save()

                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    price=product.price  # ✅ tambahkan ini agar tidak error
                )

            # Kosongkan keranjang
            try:
                cart = Cart.objects.get(user=user)
                cart.items.all().delete()
            except Cart.DoesNotExist:
                pass  # Jika tidak ada cart, tidak masalah

            return JsonResponse({"success": True, "order_id": order_id})

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"success": False, "message": str(e)}, status=500)

    return JsonResponse({"success": False, "message": "Method not allowed"}, status=405)


@csrf_exempt
def save_order(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            user_data = data.get("user")
            order_data = data.get("order")
            items_data = data.get("items")

            if not order_data or not items_data:
                return JsonResponse({"error": "Incomplete data"}, status=400)

            user = None
            if user_data:
                user = User.objects.filter(username=user_data.get("username")).first()

            order = Order.objects.create(
                user=user,
                order_id=order_data["orderNumber"],
                payment_method=order_data["paymentMethod"],
                payment_status="pending",
                shipping_address=order_data["shippingAddress"],
                delivery_estimate=order_data.get("deliveryEstimate", ""),
                total_price=Decimal(order_data["finalTotalPrice"]),
            )

            for item in items_data:
                product = Product.objects.filter(id=item["product_id"]).first()
                if product:
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=item["quantity"],
                        price=Decimal(item["price"])
                    )

            return JsonResponse({"success": True, "order_id": order.order_id})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)

#Login, Regis,dan OTP
@csrf_exempt # For development, remove and implement proper CSRF for production AJAX
def login_user(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')

            user = authenticate(request, username=username, password=password)

            print(data)
            if user is not None:
                if user.is_active:
                    login(request, user)

                    if user.is_superuser:
                        role = 'admin'
                    else:
                        role = 'user'

                    return JsonResponse({
                        'success': True, 
                        'message': 'Login successful.',
                        'role': role,
                        'username': user.username,
                    }, 
                    status=200)
                else:
                    # If user is not active, try to resend OTP
                    send_otp_to_email(user) # Resend OTP if user tries to login but is not active
                    return JsonResponse({'success': False, 'message': 'Account not activated. An OTP has been sent to your email for verification.'}, status=403)
            else:
                return JsonResponse({'success': False, 'message': 'Invalid username or password.'}, status=401)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)


@csrf_exempt
def register_user(request):
    print(">> Method:", request.method)
    print(">> Headers:", request.headers)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            email = data.get('email')
            password = data.get('password')
            name = data.get('name')
            phone = data.get('phone') 

            if not all([username, email, password]):
                return JsonResponse({'success': False, 'message': 'All fields are required.'}, status=400)

            if User.objects.filter(username=username).exists():
                return JsonResponse({'success': False, 'message': 'Username already exists.'}, status=400)
            if User.objects.filter(email=email).exists():
                return JsonResponse({'success': False, 'message': 'Email already exists.'}, status=400)

            user = User.objects.create_user(username=username, email=email, password=password)
            user.first_name = name
            user.is_active = True
            user.save()

            # ✅ Simpan phone dan name ke Profile, bukan ke user langsung
            Profile.objects.create(user=user, phone=phone)

            send_otp_to_email(user)

            return JsonResponse({'success': True, 'message': 'Registration successful. OTP sent to your email for verification.'}, status=201)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)



@csrf_exempt # For development, remove and implement proper CSRF for production AJAX
def send_otp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email_or_username = data.get('email_or_username')

            if not email_or_username:
                return JsonResponse({'success': False, 'message': 'Email or Username is required.'}, status=400)

            user = User.objects.filter(email=email_or_username).first()
            if not user:
                user = User.objects.filter(username=email_or_username).first()

            if user:
                send_otp_to_email(user)
                return JsonResponse({'success': True, 'message': 'OTP sent to your email.'}, status=200)
            else:
                return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)


@csrf_exempt # For development, remove and implement proper CSRF for production AJAX
def verify_otp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email_or_username = data.get('email_or_username')
            otp_code = data.get('otp_code')

            if not all([email_or_username, otp_code]):
                return JsonResponse({'success': False, 'message': 'Email/Username and OTP are required.'}, status=400)

            user = User.objects.filter(email=email_or_username).first() or \
                   User.objects.filter(username=email_or_username).first()

            if not user:
                return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)

            try:
                otp_obj = OTP.objects.get(user=user)
                if otp_obj.otp_code == otp_code:
                    if otp_obj.expires_at > now():
                        user.is_active = True
                        user.save()
                        otp_obj.delete()  # Delete used OTP
                        return JsonResponse({'success': True, 'message': 'Akun berhasil diverifikasi. Silakan login.'})
                    else:
                        return JsonResponse({'success': False, 'message': 'OTP sudah kedaluwarsa. Silakan minta ulang.'}, status=400)
                else:
                    return JsonResponse({'success': False, 'message': 'OTP tidak valid.'}, status=400)
            except OTP.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'OTP tidak ditemukan. Silakan minta ulang.'}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Permintaan tidak valid (JSON error).'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Metode permintaan tidak diizinkan.'}, status=405)


def send_otp_to_email(user):
    OTP.objects.filter(user=user).delete()

    otp_obj = OTP.objects.create(user=user)
    otp_code = otp_obj.generate_otp()
    otp_obj.otp_code = otp_code
    otp_obj.save()

    subject = 'Your Terusan Minimarket OTP Verification Code'
    message = (
        f'Hi {user.username},\n\nYour OTP verification code is: {otp_code}\n\n'
        'This code will expire in 5 minutes. Do not share this code with anyone.\n\n'
        'Thank you,\nTerusan Minimarket Team'
    )
    from_email = settings.EMAIL_HOST_USER  # ✅ Gunakan email dari settings.py
    recipient_list = [user.email]

    send_mail(subject, message, from_email, recipient_list, fail_silently=False)

from django.contrib.auth import logout
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt  # Gunakan sementara, atau pastikan CSRF token dikirim
def user_logout(request):
    if request.method == 'POST':
        logout(request)
        return JsonResponse({'success': True, 'message': 'Logout berhasil'})
    return JsonResponse({'success': False, 'message': 'Hanya POST diizinkan'}, status=405)

def get_current_user(request):
    if request.user.is_authenticated:
        addresses = Address.objects.filter(user=request.user).values(
            'id', 'nama_alamat', 'receiver_name', 'receiver_phone',
            'full_address', 'city', 'kabupaten', 'province',
            'postal_code', 'rt_rw', 'location_details'
        )
        return JsonResponse({
            'is_authenticated': True,
            'username': request.user.username,
            'email': request.user.email,
            'addresses': list(addresses)
        })
    return JsonResponse({'is_authenticated': False})


@csrf_exempt
@login_required
def add_to_cart(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            product_id = data.get("product_id")
            quantity = int(data.get("quantity", 1))

            product = Product.objects.get(id=product_id)

            cart, _ = Cart.objects.get_or_create(user=request.user)
            cart_item, created = CartItem.objects.get_or_create(cart=cart, product=product)

            if not created:
                if cart_item.quantity + quantity <= product.stock:
                    cart_item.quantity += quantity
                else:
                    return JsonResponse({"success": False, "message": "Stok tidak cukup."}, status=400)
            else:
                cart_item.quantity = quantity

            cart_item.save()
            return JsonResponse({"success": True, "message": "Produk ditambahkan ke keranjang."})

        except Product.DoesNotExist:
            return JsonResponse({"success": False, "message": "Produk tidak ditemukan."}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)

    return JsonResponse({"success": False, "message": "Metode tidak diperbolehkan"}, status=405)


@login_required
def get_cart_items(request):
    try:
        cart = Cart.objects.get(user=request.user)
        items = cart.items.select_related("product")

        data = [
            {
                "product_id": item.product.id,
                "product_name": item.product.name,
                "price": str(item.product.price),
                "quantity": item.quantity,
                "image": item.product.image.url if item.product.image else "/static/images/default-product.png"
            }
            for item in items
        ]
        return JsonResponse({"items": data})
    except Cart.DoesNotExist:
        return JsonResponse({"items": []})


@csrf_exempt
@login_required
def remove_cart_item(request):
    try:
        data = json.loads(request.body)
        product_id = data.get("product_id")
        if not product_id:
            return JsonResponse({"error": "Product ID is required"}, status=400)

        product = Product.objects.get(id=product_id)
        cart = Cart.objects.get(user=request.user)
        CartItem.objects.filter(cart=cart, product=product).delete()

        return JsonResponse({"message": "Item removed"})
    except Product.DoesNotExist:
        return JsonResponse({"error": "Product not found"}, status=404)
    except Cart.DoesNotExist:
        return JsonResponse({"error": "Cart not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



@csrf_exempt
@login_required
def update_cart_item(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        data = json.loads(request.body)
        product_id = data.get("product_id")
        action = data.get("action")

        if not product_id or action not in ["increase", "decrease"]:
            return JsonResponse({"error": "Data tidak valid"}, status=400)

        cart, _ = Cart.objects.get_or_create(user=request.user)
        product = get_object_or_404(Product, id=product_id)

        cart_item, _ = CartItem.objects.get_or_create(cart=cart, product=product)

        if action == "increase":
            if cart_item.quantity < product.stock:
                cart_item.quantity += 1
                cart_item.save()
            else:
                return JsonResponse({"error": "Stok tidak mencukupi"}, status=400)
        elif action == "decrease":
            cart_item.quantity -= 1
            if cart_item.quantity <= 0:
                cart_item.delete()
            else:
                cart_item.save()

        return JsonResponse({"message": "Keranjang berhasil diperbarui"})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@login_required
def clear_cart(request):
    try:
        cart = Cart.objects.get(user=request.user)
        cart.items.all().delete()  # gunakan related_name yang benar
        return JsonResponse({"message": "Keranjang dikosongkan."})
    except Cart.DoesNotExist:
        return JsonResponse({"message": "Keranjang sudah kosong."})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    


@login_required
def get_cart_count(request):
    try:
        cart, _ = Cart.objects.get_or_create(user=request.user)
        total_items = sum(item.quantity for item in cart.items.all())
        return JsonResponse({'success': True, 'total_items': total_items})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@csrf_exempt  # Only needed if you're not sending CSRF header
@login_required
def add_product(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            category_id = request.POST.get('category')
            stock = int(request.POST.get('stock', 0))
            price = float(request.POST.get('price', 0))
            description = request.POST.get('description', '')
            image = request.FILES.get('image')

            if not name or not price:
                return JsonResponse({'success': False, 'message': 'Name and price are required.'}, status=400)

            category = Category.objects.filter(id=category_id).first()
            if not category:
                return JsonResponse({'success': False, 'message': 'Category not found.'}, status=400)

            product = Product.objects.create(
                name=name,
                category=category,
                stock=stock,
                price=price,
                description=description,
                image=image
            )

            return JsonResponse({'success': True, 'message': 'Product added successfully.', 'product_id': product.id}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

def get_product_list(request):
    products = Product.objects.select_related('category').all()

    data = []
    for product in products:
        data.append({
            'id': product.id,
            'name': product.name,
            'category': product.category.name if product.category else 'Uncategorized',
            'price': float(product.price),
            'stock': product.stock,
            'image': product.image.url if product.image else '',
            'description': product.description,
        })

    return JsonResponse({'products': data})


@csrf_exempt  # Optional if CSRF token is passed from frontend
@login_required
def add_category(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()

            if not name:
                return JsonResponse({'success': False, 'message': 'Category name is required.'}, status=400)

            # Check if it already exists
            if Category.objects.filter(name__iexact=name).exists():
                return JsonResponse({'success': False, 'message': 'Category already exists.'}, status=400)

            category = Category.objects.create(name=name)

            return JsonResponse({'success': True, 'message': 'Category added.', 'id': category.id}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

def get_category_list(request):
    categories = (
        Category.objects
        .annotate(product_count=Count('product'))
        .values('id', 'name', 'product_count')
    )
    
    return JsonResponse({'categories': list(categories)}, safe=False)


def get_order_list(request):
    orders = Order.objects.values(
        'id',
        'order_id',
        'user__username',
        'payment_method',
        'payment_status',
        'shipping_address',
        'delivery_estimate',
        'total_price',
        'created_at'
    ).order_by('-created_at')

    return JsonResponse({'orders': list(orders)}, safe=False)


def get_user_list(request):
    users = User.objects.values(
        'id',
        'username',
        'email',
        'first_name',
        'last_name',
        'is_active',
        'is_staff',
        'date_joined'
    ).order_by('-date_joined')

    return JsonResponse({'users': list(users)}, safe=False)

@login_required
def profile_address_view(request):
    return render(request, 'profile_address.html')

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def profile_view(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)

    if request.method == 'GET':
        serializer = ProfileSerializer(profile)
        address_qs = Address.objects.filter(user=request.user).first()

        address_data = {
            "receiver_name": address_qs.receiver_name if address_qs else "",
            "receiver_phone": address_qs.receiver_phone if address_qs else "",
            "full_address": address_qs.full_address if address_qs else "",
            "rt_rw": address_qs.rt_rw if address_qs else "",
            "province": address_qs.province if address_qs else "",
            "kabupaten": address_qs.kabupaten if address_qs else "",
            "city": address_qs.city if address_qs else "",
            "postal_code": address_qs.postal_code if address_qs else "",
            "location_details": address_qs.location_details if address_qs else "",
            "nama_alamat": address_qs.nama_alamat if address_qs else "",
        }

        return Response({'success': True, **serializer.data, **address_data})

    elif request.method == 'POST':
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()

            if request.data.get("full_address"):
                address, _ = Address.objects.get_or_create(user=request.user)
                for field in [
                    "receiver_name", "receiver_phone", "full_address", "rt_rw",
                    "province", "kabupaten", "city", "postal_code", "location_details", "nama_alamat"
                ]:
                    setattr(address, field, request.data.get(field, ""))
                address.save()

            return Response({'success': True, 'message': 'Profil berhasil diperbarui'})
        return Response({'success': False, 'errors': serializer.errors}, status=400)

from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

@csrf_exempt
@login_required
def change_password(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        current = data.get('current_password')
        new = data.get('new_password')

        if not request.user.check_password(current):
            return JsonResponse({'success': False, 'message': 'Kata sandi lama salah'}, status=400)

        request.user.set_password(new)
        request.user.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'message': 'Metode tidak diizinkan'}, status=405)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Address
from django.shortcuts import get_object_or_404

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_address(request):
    data = request.data
    user = request.user

    required_fields = ['receiver_name', 'receiver_phone', 'full_address', 'rt_rw', 'postal_code']
    if not all(field in data and data[field] for field in required_fields):
        return Response({'success': False, 'message': 'Mohon lengkapi semua bidang'}, status=400)

    address = Address.objects.create(
        user=user,
        receiver_name=data['receiver_name'],
        receiver_phone=data['receiver_phone'],
        full_address=data['full_address'],
        rt_rw=data['rt_rw'],
        postal_code=data['postal_code'],
    )
    return Response({'success': True, 'message': 'Alamat berhasil ditambahkan', 'id': address.id})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_addresses(request):
    user = request.user
    addresses = Address.objects.filter(user=user).values(
        'id',
        'receiver_name',
        'receiver_phone',
        'full_address',
        'rt_rw',
        'postal_code',
    )
    return Response({'addresses': list(addresses)})

@api_view(["GET", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def address_detail(request, address_id):
    user = request.user
    address = get_object_or_404(Address, id=address_id, user=user)

    if request.method == "GET":
        return Response({
            'id': address.id,
            'receiver_name': address.receiver_name,
            'receiver_phone': address.receiver_phone,
            'full_address': address.full_address,
            'rt_rw': address.rt_rw,
            'postal_code': address.postal_code,
        })

    elif request.method == "PUT":
        data = request.data
        address.receiver_name = data.get('receiver_name', address.receiver_name)
        address.receiver_phone = data.get('receiver_phone', address.receiver_phone)
        address.full_address = data.get('full_address', address.full_address)
        address.rt_rw = data.get('rt_rw', address.rt_rw)
        address.postal_code = data.get('postal_code', address.postal_code)
        address.save()
        return Response({'success': True, 'message': 'Alamat berhasil diperbarui'})

    elif request.method == "DELETE":
        address.delete()
        return Response({'success': True, 'message': 'Alamat berhasil dihapus'})


# ======================= Analytics ============================
def weekly_sales_data(request):
    today = now().date()
    seven_days_ago = today - timedelta(days=6)

    # List tanggal dari 7 hari terakhir
    dates = [seven_days_ago + timedelta(days=i) for i in range(7)]
    labels = [date.strftime("%d %b") for date in dates]  # e.g. "15 Jul"
    data = []

    for date in dates:
        total = (
            Order.objects
            .filter(created_at__date=date, payment_status='settlement')
            .aggregate(total=Sum('total_price'))['total'] or 0
        )
        data.append(float(total))

    return JsonResponse({'labels': labels, 'data': data})


# === Produk Terlaris Tetap ===
def top_products_data(request):
    top_items = (
        OrderItem.objects
        .filter(order__payment_status='settlement')
        .values('product__name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:10]
    )

    labels = [item['product__name'] for item in top_items]
    data = [item['total_sold'] for item in top_items]

    return JsonResponse({'labels': labels, 'data': data})

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.timezone import now, timedelta
from .models import Order

def export_sales_report_excel(request):
    # Ambil order 30 hari terakhir
    today = now().date()
    start_date = today - timedelta(days=30)

    orders = Order.objects.filter(created_at__date__gte=start_date, payment_status='settlement')

    # Buat workbook dan worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    # Header kolom
    headers = ['Tanggal', 'ID Pesanan', 'Nama User', 'Metode Pembayaran', 'Total Harga']
    ws.append(headers)

    # Isi data
    for order in orders:
        ws.append([
            order.created_at.strftime('%d-%m-%Y'),
            order.order_id,
            order.user.username if order.user else 'Guest',
            order.payment_method,
            float(order.total_price)
        ])

    # Set lebar kolom otomatis
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Kembalikan sebagai response excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=LaporanPenjualan_{today}.xlsx'
    wb.save(response)
    return response

from django.http import JsonResponse, HttpResponseNotAllowed, HttpResponseBadRequest

def order_detail(request, pk):
    try:
        order = Order.objects.select_related("user").get(pk=pk)
        return JsonResponse({
            "order_id": order.order_id,
            "user__username": order.user.username if order.user else None,
            "created_at": order.created_at.isoformat(),
            "total_price": float(order.total_price),
            "payment_status": order.payment_status,
            "shipping_address": order.shipping_address,
        })
    except Order.DoesNotExist:
        return JsonResponse({"error": "Order tidak ditemukan"}, status=404)

def cancel_order(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(['POST'])

    try:
        order = Order.objects.get(pk=pk)
        if order.payment_status == 'settlement':
            return JsonResponse({"error": "Pesanan sudah diselesaikan, tidak bisa dibatalkan."}, status=400)

        order.payment_status = 'cancelled'
        order.save()
        return JsonResponse({"message": "Pesanan dibatalkan."})

    except Order.DoesNotExist:
        return JsonResponse({"error": "Pesanan tidak ditemukan"}, status=404)

from django.contrib.auth.models import User
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator

@csrf_exempt
@user_passes_test(lambda u: u.is_superuser)  # hanya admin/superuser yang bisa hapus
def delete_user(request, user_id):
    if request.method == 'DELETE':
        try:
            user = User.objects.get(id=user_id)
            if user.is_superuser:
                return JsonResponse({"error": "Tidak dapat menghapus superuser."}, status=403)
            user.delete()
            return JsonResponse({"message": "User berhasil dihapus."}, status=200)
        except User.DoesNotExist:
            return JsonResponse({"error": "User tidak ditemukan."}, status=404)
    return JsonResponse({"error": "Metode tidak diizinkan."}, status=405)
